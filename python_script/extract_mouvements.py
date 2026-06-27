"""Consolide les onglets de mouvements (colonnes P:Z) de Pool LT.xlsm en une seule liste."""
import openpyxl
from openpyxl import Workbook

SOURCE = "../excel/Pool LT.xlsm"
OUTPUT = "../excel/Mouvements_consolides.xlsx"

POOLERS = [
    "Vincent", "Sebastien_FAU", "Jerome", "Sebastien_STL",
    "David", "Steve", "Paule", "Nicolas",
]

HEADERS = [
    "Pooler",
    "Type",
    "Joueur acquis/activé",
    "Choix acquis",
    "Annee choix acquis",
    "Joueur cede/desactive",
    "Choix cede",
    "Annee choix cede",
    "Echange Pooler",
    "Choix Pooler",
    "Date",
    "Jusqu'a",
]

def main():
    wb_src = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Mouvements"
    ws_out.append(HEADERS)

    total = 0
    for pooler in POOLERS:
        ws = wb_src[pooler]
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=16, max_col=26, values_only=True):
            if all(v is None for v in row):
                continue
            ws_out.append([pooler, *row])
            total += 1

    wb_src.close()
    wb_out.save(OUTPUT)
    print(f"{total} lignes ecrites dans {OUTPUT}")

if __name__ == "__main__":
    main()
