"""Importe les rankings de prospects (DraftCenter) depuis un classeur Excel multi-onglets.

Un onglet par source ; le nom de l'onglet DOIT être l'une des clés de SOURCES_VALIDES
(ex: "elite_prospects", "tsn_button", "mckeens"...).

Colonnes attendues dans chaque onglet (ligne d'en-tête) — deux formats acceptés :
    1) rank, first_name, last_name, position, team, gp, g, a, p/tp, pim
    2) rank, name, team, league, gp, g, a, p/tp, pim
       où "name" = "Prénom Nom (POS)" (espace insécable + parenthèse, parfois suivi
       de texte parasite comme "Verified player") et "league" peut contenir un préfixe
       pays collé au code de ligue (ex: "USA flagNCAA" → "NCAA").
(gp/g/a/p/pim optionnelles — laisser vide si la source ne les fournit pas)

Usage:
    python import_draft_prospects.py --year 2026 --excel source/draft_prospects_2026.xlsx [--url https://...]

--url s'applique comme source_url par défaut pour tous les onglets importés dans cette
exécution (attribution). Pour des URLs différentes par source, modifier ensuite à la main
dans /admin/draft-center.
"""
import argparse
import os
import re
import sys

import openpyxl
from dotenv import load_dotenv
from supabase import create_client
from unidecode import unidecode

sys.stdout.reconfigure(encoding='utf-8')
load_dotenv()

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY')

SOURCES_VALIDES = {
    'elite_prospects', 'tsn_button', 'tsn_peters', 'mckeens', 'thn_ferrari', 'thn_kennedy',
    'daily_faceoff', 'flohockey_peters', 'central_scouting_na', 'central_scouting_eu',
    'draft_prospects_hockey', 'sportsnet_cosentino', 'sportsnet_bukala',
    'smaht_scouting', 'dobber_prospects', 'hpr_malloy',
}


def normaliser_nom(nom):
    return unidecode(str(nom)).lower().strip().replace('-', ' ')


def to_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def parser_name(raw):
    """"Gavin McKenna\xa0(LW)Verified player" -> ("Gavin", "McKenna", "LW")"""
    texte = str(raw or '').replace('\xa0', ' ').strip()
    m = re.match(r'^(.*?)\s*\(([^)]+)\)', texte)
    if m:
        nom_complet, position = m.group(1).strip(), m.group(2).strip()
    else:
        nom_complet, position = texte, None
    tokens = nom_complet.split()
    if not tokens:
        return None, None, position
    prenom = tokens[0]
    nom = ' '.join(tokens[1:]) if len(tokens) > 1 else tokens[0]
    return prenom, nom, position


def nettoyer_league(raw):
    """"USA flagNCAA" -> "NCAA" """
    texte = str(raw or '').strip()
    return re.sub(r'^.*?flag', '', texte).strip() or None


def lire_onglet(ws):
    headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in raw):
            continue
        rows.append(dict(zip(headers, raw)))
    return rows


def importer(annee, excel_path, source_url):
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    print(f'[INFO] Chargement des prospects {annee} existants...')
    existants = (
        supabase.table('draft_prospects')
        .select('id, first_name, last_name')
        .eq('draft_year', annee)
        .execute()
        .data
    )
    existing_map = {
        (normaliser_nom(p['first_name']), normaliser_nom(p['last_name'])): p['id']
        for p in existants
    }
    print(f'[INFO] {len(existing_map)} prospects en base pour {annee}')

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    total_crees = 0
    total_rangs = 0

    for sheet_name in wb.sheetnames:
        if sheet_name not in SOURCES_VALIDES:
            print(f'[SKIP] Onglet "{sheet_name}" ignoré — pas une source reconnue ({sorted(SOURCES_VALIDES)})')
            continue

        rows = lire_onglet(wb[sheet_name])
        print(f'[INFO] Onglet "{sheet_name}": {len(rows)} lignes')

        rankings_a_inserer = []

        for row in rows:
            rang = to_int(row.get('rank'))

            if row.get('first_name') or row.get('last_name'):
                prenom = str(row.get('first_name') or '').strip()
                nom = str(row.get('last_name') or '').strip()
                position = str(row.get('position') or '').strip() or None
            else:
                prenom, nom, position = parser_name(row.get('name'))
                prenom = (prenom or '').strip()
                nom = (nom or '').strip()

            if not prenom or not nom or rang is None:
                print(f'  [SKIP] ligne incomplète: {row}')
                continue

            team_brut = str(row.get('team') or '').strip() or None
            league = nettoyer_league(row.get('league')) if row.get('league') else None
            if team_brut and league:
                team = f'{team_brut}, {league}'
            else:
                team = team_brut or league

            stats = {
                'games_played': to_int(row.get('gp')),
                'goals':        to_int(row.get('g')),
                'assists':      to_int(row.get('a')),
                'points':       to_int(row.get('p')) if row.get('p') is not None else to_int(row.get('tp')),
                'pim':          to_int(row.get('pim')),
            }
            has_stats = any(v is not None for v in stats.values())

            key = (normaliser_nom(prenom), normaliser_nom(nom))
            prospect_id = existing_map.get(key)

            if not prospect_id:
                inserted = (
                    supabase.table('draft_prospects')
                    .insert({
                        'draft_year': annee,
                        'first_name': prenom,
                        'last_name': nom,
                        'position': position,
                        'team': team,
                        **stats,
                    })
                    .execute()
                    .data
                )
                prospect_id = inserted[0]['id']
                existing_map[key] = prospect_id
                total_crees += 1
            elif has_stats or position or team:
                # Met à jour les infos bio/stats si cet onglet en fournit
                update = {k: v for k, v in {'position': position, 'team': team, **stats}.items() if v is not None}
                if update:
                    supabase.table('draft_prospects').update(update).eq('id', prospect_id).execute()

            rankings_a_inserer.append({
                'prospect_id': prospect_id,
                'source': sheet_name,
                'source_url': source_url,
                'rank': rang,
            })

        if rankings_a_inserer:
            supabase.table('draft_prospect_rankings').upsert(
                rankings_a_inserer, on_conflict='prospect_id,source'
            ).execute()
        total_rangs += len(rankings_a_inserer)

    wb.close()
    print(f'[INFO] {total_crees} nouveaux prospects créés, {total_rangs} rangs importés/mis à jour.')
    print('[INFO] Terminé.')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--year', type=int, required=True)
    parser.add_argument('--excel', required=True)
    parser.add_argument('--url', default=None)
    args = parser.parse_args()
    importer(args.year, args.excel, args.url)
