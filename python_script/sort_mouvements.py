"""Trie Mouvements_consolides.xlsx par ordre chronologique.

Certaines lignes n'ont pas de Date renseignee (changements "permanents" sans
fenetre temporaire, ex: echanges, BLT). Pour ces lignes, on reutilise la
derniere Date connue du meme pooler (le tableau source est deja chronologique
par onglet), sinon la prochaine Date connue si aucune ne precede. Une colonne
"Date estimee" indique quand la date a ete deduite plutot que lue directement.
"""
import openpyxl
from openpyxl import Workbook

INPUT = "../excel/Mouvements_consolides.xlsx"
OUTPUT = "../excel/Mouvements_consolides.xlsx"

def main():
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

    date_col = headers.index("Date")

    # Forward-fill puis backward-fill la date par pooler, en respectant l'ordre original.
    by_pooler = {}
    for idx, row in enumerate(rows):
        by_pooler.setdefault(row[0], []).append(idx)

    resolved_date = [None] * len(rows)
    estimee = [False] * len(rows)

    for pooler, idxs in by_pooler.items():
        last_known = None
        for i in idxs:
            d = rows[i][date_col]
            if d is not None:
                last_known = d
                resolved_date[i] = d
            else:
                resolved_date[i] = last_known
                estimee[i] = True
        # backward-fill pour les lignes en tete de liste sans date precedente
        next_known = None
        for i in reversed(idxs):
            if rows[i][date_col] is not None:
                next_known = rows[i][date_col]
            elif resolved_date[i] is None:
                resolved_date[i] = next_known

    for i, row in enumerate(rows):
        row.append(resolved_date[i])
        row.append("Oui" if estimee[i] else "Non")

    # tri stable : date resolue, puis ordre d'origine (deja chronologique par onglet)
    indexed = list(enumerate(rows))
    indexed.sort(key=lambda t: (t[1][-2] is None, t[1][-2], t[0]))
    sorted_rows = [r for _, r in indexed]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Mouvements"
    ws_out.append(headers + ["Date tri", "Date estimee"])
    for row in sorted_rows:
        ws_out.append(row)

    wb_out.save(OUTPUT)
    print(f"{len(sorted_rows)} lignes triees et ecrites dans {OUTPUT}")

if __name__ == "__main__":
    main()
