"""
Reconstruit l'historique de roster de la saison régulière active (pooler_rosters,
roster_change_log) à partir du suivi personnel de David, excel/Mouvements_consolides.xlsx
(hors dépôt git — voir CLAUDE.md section 1 et SUIVI_PROJET.md 2026-07-31/2026-08-03).

Contexte : buildStandings() (app/lib/standings.ts) attribue les points d'un joueur selon
la fenêtre added_at→removed_at de sa ligne pooler_rosters et les transitions de statut
journalisées dans roster_change_log. Cette histoire fine n'a jamais été saisie au fil de
l'eau pour 2025-26 — seule une poignée de corrections ponctuelles existe via
/admin/historique. Ce script rejoue les ~270 mouvements du fichier Excel pour reconstruire
cette histoire automatiquement, plutôt que de la ressaisir un par un dans l'UI.

Mécanique (indépendante du texte libre de la colonne "Type", jugé peu fiable — voir le
plan de session) : on simule chronologiquement, pour chaque joueur touché, qui le détient
et à quel statut, en se basant uniquement sur les colonnes "Statut joueur acquis/cédé" et
la présence de "Echange Pooler". Le côté acquis regarde l'état simulé courant du joueur
pour décider s'il s'agit d'un changement de statut interne, d'un vrai transfert, ou d'un
ajout neuf. Le côté cédé quitte réellement le pool si son statut est "Ballotage", part
vers "Echange Pooler" si rempli, sinon reste chez le même pooler à son nouveau statut.

Portée : uniquement pooler_rosters/roster_change_log. Les colonnes de choix de repêchage
(Choix acquis/cédé, Choix Pooler) sont seulement listées dans le rapport, jamais rejouées
dans pool_draft_picks (décidé avec David le 2026-08-03).

Usage:
    python import_mouvements_excel.py            # dry-run : rapport seulement, rien écrit
    python import_mouvements_excel.py --apply     # exécute réellement (confirmation "oui")
"""

import os
import sys
import re
import difflib
from collections import defaultdict
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

from dotenv import load_dotenv
from supabase import create_client
import openpyxl
from unidecode import unidecode

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.normpath(os.path.join(BASE_DIR, '..', 'excel', 'Mouvements_consolides.xlsx'))

POOLER_NAME_MAP = {
    'vincent': 'Vincent',
    'paule': 'Paule',
    'nicolas': 'Nicolas',
    'steve': 'Steve',
    'david': 'David',
    'jerome': 'Jérôme',
    'sebastien_fau': 'Sébastien F.',
    'sebastien_stl': 'Sébastien S.',
    'sebastien s.': 'Sébastien S.',
}

STATUS_MAP = {
    'actif': 'actif',
    'reserviste': 'reserviste',
    'recrue': 'recrue',
    'ir': 'ltir',
    'ballotage': 'BALLOTAGE',
}

ACTIVE_LIMITS = {'forward': 12, 'defense': 6, 'goalie': 2}


class Tee:
    """Duplique stdout vers un fichier de log, comme les autres scripts du pipeline."""
    def __init__(self, path):
        self.file = open(path, 'w', encoding='utf-8')
        self.stdout = sys.stdout

    def write(self, data):
        self.stdout.write(data)
        self.file.write(data)

    def flush(self):
        self.stdout.flush()
        self.file.flush()


def norm(s):
    return unidecode(str(s or '')).lower().strip()


def connect():
    load_dotenv(os.path.join(BASE_DIR, '.env.staging'), override=True)
    url = os.environ['SUPABASE_URL']
    key = os.environ['SUPABASE_SERVICE_KEY']
    return create_client(url, key)


def fetch_all(db, table, select, filters=None):
    rows, offset, page = [], 0, 1000
    filters = filters or {}
    while True:
        q = db.table(table).select(select)
        for k, v in filters.items():
            q = q.eq(k, v)
        r = q.range(offset, offset + page - 1).execute()
        rows.extend(r.data)
        if len(r.data) < page:
            break
        offset += page
    return rows


def get_active_season(db):
    r = db.table('pool_seasons').select('id, season, saison_start_date, saison_end_date') \
        .eq('is_active', True).eq('is_playoff', False).single().execute()
    if not r.data:
        raise SystemExit('[ERREUR] Aucune saison régulière active en staging.')
    return r.data


def get_player_bucket(position):
    pos = (position or '').upper()
    if 'G' in pos:
        return 'goalie'
    if 'D' in pos:
        return 'defense'
    return 'forward'


# ---------------------------------------------------------------------------
# Lecture du fichier Excel
# ---------------------------------------------------------------------------

def load_excel_rows():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Mouvements']
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s[:10] if s else None


def parse_player_name(cell):
    if not cell:
        return None
    s = str(cell).strip()
    if not s:
        return None
    # format attendu : "Prénom Nom - EQUIPE - POS,POS2" — on ne garde que le nom
    return s.split(' - ')[0].strip()


def map_pooler(cell):
    if not cell:
        return None
    key = norm(cell)
    return POOLER_NAME_MAP.get(key)


def map_status(cell):
    if not cell:
        return None
    key = norm(cell)
    return STATUS_MAP.get(key)


# ---------------------------------------------------------------------------
# Résolution des noms de joueurs
# ---------------------------------------------------------------------------

class PlayerIndex:
    def __init__(self, players):
        self.by_id = {p['id']: p for p in players}
        self.by_key = defaultdict(list)
        for p in players:
            key = norm(f"{p['first_name']} {p['last_name']}")
            self.by_key[key].append(p['id'])
        self.all_names = list(self.by_key.keys())

    def resolve(self, name):
        """Retourne (player_id, note) — player_id est None si non résolu/ambigu."""
        key = norm(name)
        candidates = self.by_key.get(key)
        if candidates and len(candidates) == 1:
            return candidates[0], None
        if candidates and len(candidates) > 1:
            return None, f"Nom ambigu ({len(candidates)} joueurs) : {name}"
        suggestion = difflib.get_close_matches(key, self.all_names, n=1, cutoff=0.8)
        if suggestion:
            return None, f"Nom introuvable : {name!r} — vouliez-vous dire {self.by_id[self.by_key[suggestion[0]][0]]['first_name']} {self.by_id[self.by_key[suggestion[0]][0]]['last_name']!r} ?"
        return None, f"Nom introuvable : {name!r} (aucune suggestion)"


# ---------------------------------------------------------------------------
# Rapport
# ---------------------------------------------------------------------------

class Report:
    def __init__(self):
        self.unresolved_names = []       # (row_idx, date, pooler, side, raw_name, note)
        self.bootstraps = []             # (pooler_name, player_id, date, note)
        self.echange_mismatches = []     # (row_idx, date, pooler, player_id, declared, actual)
        self.anomalies = []              # (row_idx, date, pooler, message)
        self.pick_rows = []              # (row_idx, date, pooler, description)
        self.legality_violations = []    # (date, pooler, message)
        self.rows_processed = 0
        self.players_touched = set()
        self.poolers_touched = set()


# ---------------------------------------------------------------------------
# Simulation
# ---------------------------------------------------------------------------

class Simulation:
    def __init__(self, season_start_ts):
        self.season_start_ts = season_start_ts
        self.current_owner = {}                      # player_id -> pooler_name | None
        self.rows_by_pair = defaultdict(list)         # (pooler_name, player_id) -> [row dicts]

    def get_open_row(self, pooler, player_id):
        lst = self.rows_by_pair.get((pooler, player_id))
        if lst and lst[-1]['removed_at'] is None:
            return lst[-1]
        return None

    def ensure_owner(self, player_id, pooler, date, report, reason):
        """Si le joueur n'a jamais été vu, on suppose qu'il était déjà chez `pooler`
        depuis le début de la saison, au statut 'actif' (meilleure estimation générique —
        la très grande majorité des mouvements du fichier partent d'un joueur actif qu'on
        rétrograde). Marqué dans le rapport pour vérification manuelle."""
        if self.current_owner.get(player_id) == pooler and self.get_open_row(pooler, player_id):
            return
        if self.get_open_row(pooler, player_id) is None:
            row = {
                'added_at': self.season_start_ts,
                'removed_at': None,
                'player_type': 'actif',
                'transitions': [(self.season_start_ts, None, 'actif')],
                'bootstrap': True,
            }
            self.rows_by_pair[(pooler, player_id)].append(row)
            report.bootstraps.append((pooler, player_id, date, reason))
        self.current_owner[player_id] = pooler

    def close_row(self, pooler, player_id, date, report):
        row = self.get_open_row(pooler, player_id)
        if row is None:
            report.anomalies.append((None, date, pooler,
                f"Retrait impossible : aucune ligne ouverte pour joueur {player_id} chez {pooler}"))
            return
        row['removed_at'] = date
        if self.current_owner.get(player_id) == pooler:
            self.current_owner[player_id] = None

    def open_row(self, pooler, player_id, date, ptype):
        row = {
            'added_at': date,
            'removed_at': None,
            'player_type': ptype,
            'transitions': [(date, None, ptype)],
            'bootstrap': False,
        }
        self.rows_by_pair[(pooler, player_id)].append(row)
        self.current_owner[player_id] = pooler

    def change_type(self, pooler, player_id, date, new_type, report, reason):
        self.ensure_owner(player_id, pooler, date, report, reason)
        row = self.get_open_row(pooler, player_id)
        old = row['player_type']
        if old == new_type:
            return
        row['player_type'] = new_type
        row['transitions'].append((date, old, new_type))

    def process_acquis(self, pooler, player_id, date, new_type, echange_pooler, row_idx, report):
        cur = self.current_owner.get(player_id)
        if echange_pooler and cur is not None and cur != echange_pooler and cur != pooler:
            report.echange_mismatches.append((row_idx, date, pooler, player_id, echange_pooler, cur))
        if cur is None and echange_pooler and echange_pooler != pooler:
            self.ensure_owner(player_id, echange_pooler, date, report,
                               f"Origine déduite via 'Echange Pooler' (ligne {row_idx})")
            cur = echange_pooler
        if cur == pooler:
            self.change_type(pooler, player_id, date, new_type, report, f"ligne {row_idx}")
        elif cur is not None:
            self.close_row(cur, player_id, date, report)
            self.open_row(pooler, player_id, date, new_type)
        else:
            self.open_row(pooler, player_id, date, new_type)

    def process_cede(self, pooler, player_id, date, new_type_raw, echange_pooler, row_idx, report):
        self.ensure_owner(player_id, pooler, date, report, f"Première mention (côté cédé, ligne {row_idx})")
        if new_type_raw == 'BALLOTAGE':
            self.close_row(pooler, player_id, date, report)
        elif echange_pooler and echange_pooler != pooler:
            self.close_row(pooler, player_id, date, report)
            self.open_row(echange_pooler, player_id, date, new_type_raw or 'actif')
        elif new_type_raw:
            self.change_type(pooler, player_id, date, new_type_raw, report, f"ligne {row_idx}")
        else:
            report.anomalies.append((row_idx, date, pooler, f"Statut cédé manquant pour joueur {player_id}"))


# ---------------------------------------------------------------------------
# Légalité (best-effort, non bloquant — cap volontairement omis, voir plan)
# ---------------------------------------------------------------------------

def check_legality(sim, baseline_roster, touched_ids, positions, pooler, date, report):
    """baseline_roster: pooler_name -> {player_id: player_type} pour les joueurs JAMAIS
    touchés par le fichier (état DB courant, valide tout au long puisque non modifié).
    Les joueurs touchés sont exclus du baseline — leur présence/statut à `date` vient
    uniquement de l'état simulé (absent si pas encore acquis ou déjà cédé à `date`)."""
    roster = {pid: t for pid, t in baseline_roster.get(pooler, {}).items() if pid not in touched_ids}
    for (p, pid), lst in sim.rows_by_pair.items():
        if p != pooler:
            continue
        row = lst[-1]
        if row['added_at'] <= date and (row['removed_at'] is None or row['removed_at'] > date):
            roster[pid] = row['player_type']

    actifs = [pid for pid, t in roster.items() if t == 'actif']
    reserves = [pid for pid, t in roster.items() if t == 'reserviste']
    counts = {'forward': 0, 'defense': 0, 'goalie': 0}
    for pid in actifs:
        counts[get_player_bucket(positions.get(pid))] += 1

    msgs = []
    if counts['forward'] > ACTIVE_LIMITS['forward']:
        msgs.append(f"Trop d'attaquants actifs ({counts['forward']}/{ACTIVE_LIMITS['forward']})")
    if counts['defense'] > ACTIVE_LIMITS['defense']:
        msgs.append(f"Trop de défenseurs actifs ({counts['defense']}/{ACTIVE_LIMITS['defense']})")
    if counts['goalie'] > ACTIVE_LIMITS['goalie']:
        msgs.append(f"Trop de gardiens actifs ({counts['goalie']}/{ACTIVE_LIMITS['goalie']})")
    if len(reserves) < 2:
        msgs.append(f"Moins de 2 réservistes ({len(reserves)})")
    for m in msgs:
        report.legality_violations.append((date, pooler, m))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    apply_mode = '--apply' in sys.argv

    os.makedirs(os.path.join(BASE_DIR, 'logs'), exist_ok=True)
    log_path = os.path.join(BASE_DIR, 'logs', f'import_mouvements_{datetime.now():%Y-%m-%d_%H-%M-%S}.log')
    sys.stdout = Tee(log_path)

    print(f"[INFO] Fichier Excel : {EXCEL_PATH}")
    print(f"[INFO] Mode : {'APPLY' if apply_mode else 'DRY-RUN'}")

    db = connect()
    season = get_active_season(db)
    season_id = season['id']
    season_start_ts = f"{season['saison_start_date']}T12:00:00Z"
    print(f"[INFO] Saison active : {season['season']} (id={season_id}), début={season['saison_start_date']}")

    players = fetch_all(db, 'players', 'id, first_name, last_name, position')
    positions = {p['id']: p['position'] for p in players}
    pindex = PlayerIndex(players)
    print(f"[INFO] {len(players)} joueurs chargés depuis staging.")

    poolers_rows = db.table('poolers').select('id, name').execute().data
    pooler_name_to_id = {p['name']: p['id'] for p in poolers_rows}

    # Baseline : état courant complet, pour les joueurs qui ne seront jamais touchés par
    # le fichier Excel (sert uniquement au calcul de légalité — non modifié par le script).
    current_rosters = fetch_all(db, 'pooler_rosters',
                                 'pooler_id, player_id, player_type, added_at, removed_at',
                                 filters={'pool_season_id': season_id})
    id_to_pooler_name = {p['id']: p['name'] for p in poolers_rows}
    baseline_roster = defaultdict(dict)
    for r in current_rosters:
        if r['removed_at'] is None:
            pname = id_to_pooler_name.get(r['pooler_id'])
            if pname:
                baseline_roster[pname][r['player_id']] = r['player_type']

    excel_rows = load_excel_rows()
    print(f"[INFO] {len(excel_rows)} lignes lues dans la feuille Mouvements.")

    # Le fichier n'est pas garanti parfaitement trié (ex: lignes ajoutées après le dernier
    # passage de sort_mouvements.py, sans "Date tri" — repli sur "Date"). La simulation
    # dépend d'un ordre chronologique strict, donc on trie explicitement ici (tri stable :
    # égalité de date -> ordre d'origine dans le fichier, déjà cohérent par pooler).
    indexed_rows = list(enumerate(excel_rows, start=2))  # ligne 2 = première ligne de données
    def row_date_key(item):
        _, r = item
        d = to_date_str(r.get('Date tri')) or to_date_str(r.get('Date'))
        return d or '9999-99-99'
    indexed_rows.sort(key=row_date_key)

    sim = Simulation(season_start_ts)
    report = Report()

    for idx, r in indexed_rows:
        pooler = map_pooler(r.get('Pooler'))
        if not pooler:
            report.anomalies.append((idx, None, r.get('Pooler'), f"Pooler non reconnu : {r.get('Pooler')!r}"))
            continue
        date = to_date_str(r.get('Date tri')) or to_date_str(r.get('Date'))
        if not date:
            report.anomalies.append((idx, None, pooler, "Aucune date exploitable (Date tri et Date vides)"))
            continue
        ts = f"{date}T12:00:00Z"
        echange_pooler = map_pooler(r.get('Echange Pooler')) if r.get('Echange Pooler') else None
        report.rows_processed += 1
        report.poolers_touched.add(pooler)
        if echange_pooler:
            report.poolers_touched.add(echange_pooler)

        # Choix de repêchage — hors scope, listés seulement
        if r.get('Choix acquis') or r.get('Choix cede'):
            report.pick_rows.append((idx, date, pooler,
                f"acquis={r.get('Choix acquis')}/{r.get('Annee choix acquis')} "
                f"cédé={r.get('Choix cede')}/{r.get('Annee choix cede')} "
                f"choix_pooler={r.get('Choix Pooler')}"))

        acquis_name = parse_player_name(r.get('Joueur acquis/activé'))
        if acquis_name:
            pid, note = pindex.resolve(acquis_name)
            if pid is None:
                report.unresolved_names.append((idx, date, pooler, 'acquis', acquis_name, note))
            else:
                report.players_touched.add(pid)
                stype = map_status(r.get('Statut joueur acquis'))
                if stype is None or stype == 'BALLOTAGE':
                    report.anomalies.append((idx, date, pooler,
                        f"Statut acquis invalide/manquant pour {acquis_name} : {r.get('Statut joueur acquis')!r}"))
                else:
                    sim.process_acquis(pooler, pid, ts, stype, echange_pooler, idx, report)

        cede_name = parse_player_name(r.get('Joueur cede/desactive'))
        if cede_name:
            pid, note = pindex.resolve(cede_name)
            if pid is None:
                report.unresolved_names.append((idx, date, pooler, 'cédé', cede_name, note))
            else:
                report.players_touched.add(pid)
                stype = map_status(r.get('Statut joueur cédé'))
                if stype is None:
                    report.anomalies.append((idx, date, pooler,
                        f"Statut cédé invalide/manquant pour {cede_name} : {r.get('Statut joueur cédé')!r}"))
                else:
                    sim.process_cede(pooler, pid, ts, stype, echange_pooler, idx, report)

        # Légalité — best-effort, non bloquant (cap volontairement omis)
        for p in {pooler} | ({echange_pooler} if echange_pooler else set()):
            check_legality(sim, baseline_roster, report.players_touched, positions, p, ts, report)

    # -----------------------------------------------------------------
    # Rapport
    # -----------------------------------------------------------------
    print("\n" + "=" * 70)
    print("RAPPORT")
    print("=" * 70)
    print(f"Lignes traitées      : {report.rows_processed} / {len(excel_rows)}")
    print(f"Joueurs touchés      : {len(report.players_touched)}")
    print(f"Poolers touchés      : {len(report.poolers_touched)} — {sorted(report.poolers_touched)}")

    print(f"\n--- Noms non résolus ({len(report.unresolved_names)}) ---")
    for idx, date, pooler, side, name, note in report.unresolved_names:
        print(f"  L{idx} [{date}] {pooler} ({side}) : {note}")

    print(f"\n--- Origines déduites / bootstrap ({len(report.bootstraps)}) ---")
    print("(joueur supposé déjà présent depuis le début de saison, statut de départ estimé 'actif' — à vérifier si le résultat semble faux)")
    for pooler, pid, date, reason in report.bootstraps:
        p = pindex.by_id.get(pid, {})
        print(f"  [{date}] {pooler} <- {p.get('first_name','?')} {p.get('last_name','?')} ({reason})")

    print(f"\n--- Mésaccords 'Echange Pooler' déclaré vs propriétaire simulé ({len(report.echange_mismatches)}) ---")
    for idx, date, pooler, pid, declared, actual in report.echange_mismatches:
        p = pindex.by_id.get(pid, {})
        print(f"  L{idx} [{date}] {pooler} : {p.get('first_name','?')} {p.get('last_name','?')} — Echange Pooler={declared}, propriétaire simulé={actual}")

    print(f"\n--- Anomalies diverses ({len(report.anomalies)}) ---")
    for idx, date, pooler, msg in report.anomalies:
        print(f"  L{idx} [{date}] {pooler} : {msg}")

    print(f"\n--- Lignes avec choix de repêchage, non traitées ({len(report.pick_rows)}) ---")
    for idx, date, pooler, desc in report.pick_rows:
        print(f"  L{idx} [{date}] {pooler} : {desc}")

    print(f"\n--- Violations de légalité (non bloquantes, {len(report.legality_violations)}) ---")
    print("(NOTE : le baseline des joueurs jamais touchés par le fichier vient de l'état DB")
    print(" ACTUEL, utilisé comme approximation constante sur toute la saison — un pooler")
    print(" déjà en dépassement aujourd'hui sur ses joueurs non touchés apparaîtra donc en")
    print(" dépassement à chaque ligne le concernant, même si ce n'était pas forcément vrai")
    print(" à l'époque. Signal à interpréter avec prudence, pas une liste de vrais problèmes.)")
    seen = set()
    for date, pooler, msg in report.legality_violations:
        key = (date, pooler, msg)
        if key in seen:
            continue
        seen.add(key)
        print(f"  [{date}] {pooler} : {msg}")

    # Diff de sanité : état simulé final vs état actuel réel en base, pour les joueurs touchés
    print(f"\n--- Diff de sanité (simulé final vs DB actuelle, joueurs touchés) ---")
    current_by_player = defaultdict(dict)
    for r in current_rosters:
        if r['removed_at'] is None:
            pname = id_to_pooler_name.get(r['pooler_id'])
            if pname:
                current_by_player[r['player_id']][pname] = r['player_type']

    diffs = 0
    for pid in sorted(report.players_touched):
        sim_final = {}
        for (pooler, ppid), lst in sim.rows_by_pair.items():
            if ppid != pid:
                continue
            row = lst[-1]
            if row['removed_at'] is None:
                sim_final[pooler] = row['player_type']
        db_final = current_by_player.get(pid, {})
        if sim_final != db_final:
            diffs += 1
            p = pindex.by_id.get(pid, {})
            print(f"  {p.get('first_name','?')} {p.get('last_name','?')} — simulé={sim_final} / DB actuelle={db_final}")
    print(f"  ({diffs} joueurs avec un écart sur {len(report.players_touched)} touchés)")

    print("\n" + "=" * 70)
    if not apply_mode:
        print("[DRY-RUN] Aucune écriture effectuée. Relancer avec --apply pour appliquer.")
        return

    # -----------------------------------------------------------------
    # Application réelle
    # -----------------------------------------------------------------
    confirm = input(f"\nAppliquer ces changements à la saison {season['season']} en staging ? (tapez 'oui') ")
    if confirm.strip().lower() != 'oui':
        print("[ANNULÉ] Aucune écriture effectuée.")
        return

    touched_ids = sorted(report.players_touched)
    print(f"[APPLY] Suppression de l'historique existant pour {len(touched_ids)} joueurs...")
    for pid in touched_ids:
        db.table('roster_change_log').delete().eq('pool_season_id', season_id).eq('player_id', pid).execute()
        db.table('pooler_rosters').delete().eq('pool_season_id', season_id).eq('player_id', pid).execute()

    print("[APPLY] Insertion de l'historique simulé...")
    inserted_rows, inserted_logs = 0, 0
    for (pooler, pid), lst in sim.rows_by_pair.items():
        pooler_id = pooler_name_to_id.get(pooler)
        if not pooler_id:
            print(f"[ERREUR] Pooler inconnu en base : {pooler}")
            continue
        for row in lst:
            db.table('pooler_rosters').insert({
                'pooler_id': pooler_id,
                'player_id': pid,
                'pool_season_id': season_id,
                'player_type': row['player_type'],
                'is_active': row['removed_at'] is None,
                'added_at': row['added_at'],
                'removed_at': row['removed_at'],
            }).execute()
            inserted_rows += 1
            for changed_at, old_type, new_type in row['transitions']:
                db.table('roster_change_log').insert({
                    'player_id': pid, 'pooler_id': pooler_id, 'pool_season_id': season_id,
                    'change_type': 'excel_import', 'old_type': old_type, 'new_type': new_type,
                    'changed_by': None, 'changed_at': changed_at, 'is_admin_override': True,
                }).execute()
                inserted_logs += 1
            if row['removed_at'] is not None:
                db.table('roster_change_log').insert({
                    'player_id': pid, 'pooler_id': pooler_id, 'pool_season_id': season_id,
                    'change_type': 'excel_import', 'old_type': row['player_type'], 'new_type': None,
                    'changed_by': None, 'changed_at': row['removed_at'], 'is_admin_override': True,
                }).execute()
                inserted_logs += 1

    print(f"[APPLY] {inserted_rows} lignes pooler_rosters, {inserted_logs} lignes roster_change_log insérées.")


if __name__ == '__main__':
    main()
